from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(r"C:\Users\avoba\Downloads\finviz_all_tickers.xlsx")


def clean_symbol(value: object) -> str:
    symbol = str(value or "").strip().upper()
    symbol = re.sub(r"\s+", "", symbol)
    if not symbol or symbol in {"TICKER", "SYMBOL"}:
        return ""
    return symbol


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        raise SystemExit(f"Workbook not found: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    if "Tickers" not in workbook.sheetnames:
        raise SystemExit("Workbook must include a 'Tickers' sheet")

    sheet = workbook["Tickers"]
    headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: i for i, name in enumerate(headers)}

    required = ["Ticker", "Company"]
    missing = [name for name in required if name not in index]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")

    symbols: list[str] = []
    seen: set[str] = set()
    meta: dict[str, dict[str, str]] = {}
    provider_map: dict[str, str] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        symbol = clean_symbol(row[index["Ticker"]])
        if not symbol or symbol in seen:
            continue

        seen.add(symbol)
        symbols.append(symbol)

        company = str(row[index["Company"]] or symbol).strip()
        sector = str(row[index.get("Sector", -1)] or "Unknown").strip() if "Sector" in index else "Unknown"
        industry = str(row[index.get("Industry", -1)] or "").strip() if "Industry" in index else ""
        country = str(row[index.get("Country", -1)] or "").strip() if "Country" in index else ""
        finviz_url = str(row[index.get("Finviz URL", -1)] or "").strip() if "Finviz URL" in index else ""

        meta[symbol] = {
            "name": company,
            "market": sector or "Unknown",
            "sector": sector,
            "industry": industry,
            "country": country,
            "finviz_url": finviz_url,
        }

        # Yahoo Finance accepts Finviz class/share suffixes with hyphens.
        provider_map[symbol] = symbol

    if not symbols:
        raise SystemExit("No tickers found in workbook")

    (ROOT / "config" / "tickers.csv").write_text("\n".join(symbols) + "\n", encoding="utf-8")
    (ROOT / "apps" / "web" / "data" / "universe.json").write_text(
        json.dumps({"source": "Finviz screener", "symbols": symbols}, indent=2) + "\n",
        encoding="utf-8",
    )
    (ROOT / "apps" / "web" / "data" / "symbol_meta.json").write_text(
        json.dumps(meta, indent=2) + "\n",
        encoding="utf-8",
    )
    (ROOT / "config" / "provider_map.json").write_text(
        json.dumps(provider_map, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Imported {len(symbols)} Finviz tickers from {source}")


if __name__ == "__main__":
    main()
