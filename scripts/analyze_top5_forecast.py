#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import random
import statistics
import sys
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

FEATURES = [
    "close_gt_sma200",
    "sma200_slope_20d",
    "avg_dollar_volume_20",
    "rvol_20",
    "ret_4w",
    "ret_8w",
    "ret_12w",
    "dist_52w_high",
    "atr_pct_20",
    "compression_20_100",
    "breakout_20",
    "up_down_vol_ratio_20",
    "structure_score",
]


@dataclass
class SignalRow:
    symbol: str
    symbol_name: str
    market: str
    timeframe: str
    signal: str
    candles_ago: int
    signal_price: float
    current_price: float
    ranking: int
    week: int
    label: int


@dataclass
class Bar:
    date: dt.date
    open: float
    high: float
    low: float
    close: float
    volume: float


@dataclass
class Sample:
    week: int
    symbol: str
    label: int
    ranking: int
    features: Dict[str, float]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Walk-forward analysis of Top 5 weekly winners")
    p.add_argument("--xlsx", default=r"C:\Users\avoba\OneDrive\Desktop\Top 5.xlsx")
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--asof-date", default=str(dt.date.today()))
    p.add_argument("--weeks", type=int, default=23)
    p.add_argument("--top-k", type=int, default=5)
    p.add_argument("--min-train-weeks", type=int, default=6)
    p.add_argument("--cache-dir", default="scripts/.cache/top5_forecast")
    p.add_argument("--out-dir", default="scripts/out")
    return p.parse_args()


def to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return n if math.isfinite(n) else None


def to_int(v: object) -> Optional[int]:
    n = to_float(v)
    return int(n) if n is not None else None


def read_sheet(path: Path, sheet: str, weeks: int) -> List[SignalRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows: List[SignalRow] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        symbol = str(r[0] or "").strip()
        if not symbol:
            continue
        signal = str(r[4] or "").strip().upper()
        candles = to_int(r[5])
        signal_price = to_float(r[6])
        current_price = to_float(r[7])
        ranking = to_int(r[11])
        if signal != "BUY":
            continue
        if candles is None or candles < 1 or candles > weeks:
            continue
        if signal_price is None or current_price is None or signal_price <= 0:
            continue
        if ranking is None or ranking < 1:
            continue
        rows.append(SignalRow(symbol, str(r[1] or "").strip(), str(r[2] or "").strip(), str(r[3] or "").strip(), signal, candles, signal_price, current_price, ranking, candles, 1 if ranking <= 5 else 0))
    return rows


def yahoo_candidates(raw_symbol: str) -> List[str]:
    s = raw_symbol.strip().upper()
    out = [s]
    if ":" in s:
        out.append(s.split(":", 1)[1])
    if "/" in s:
        out.append(s.replace("/", "-"))
    if "." in s:
        out.append(s.replace(".", "-"))
    seen = set()
    dedup = []
    for x in out:
        if x and x not in seen:
            dedup.append(x)
            seen.add(x)
    return dedup


def parse_yahoo_payload(payload: dict) -> List[Bar]:
    result = (payload.get("chart", {}).get("result") or [])
    if not result:
        return []
    r0 = result[0]
    timestamps = r0.get("timestamp") or []
    quote = ((r0.get("indicators") or {}).get("quote") or [{}])[0]
    opens = quote.get("open") or []
    highs = quote.get("high") or []
    lows = quote.get("low") or []
    closes = quote.get("close") or []
    volumes = quote.get("volume") or []
    n = min(len(timestamps), len(opens), len(highs), len(lows), len(closes), len(volumes))
    bars: List[Bar] = []
    for i in range(n):
        o, h, l, c, v = to_float(opens[i]), to_float(highs[i]), to_float(lows[i]), to_float(closes[i]), to_float(volumes[i])
        if None in (o, h, l, c, v):
            continue
        d = dt.datetime.fromtimestamp(int(timestamps[i]), dt.timezone.utc).date()
        bars.append(Bar(d, o or 0.0, h or 0.0, l or 0.0, c or 0.0, v or 0.0))
    bars.sort(key=lambda b: b.date)
    return bars


def fetch_yahoo_bars(symbol: str, start: dt.date, end: dt.date, cache_dir: Path) -> List[Bar]:
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{symbol.replace(':', '_')}.json"
    if cache_file.exists():
        try:
            return parse_yahoo_payload(json.loads(cache_file.read_text(encoding="utf-8")))
        except Exception:
            pass

    p1 = int(dt.datetime.combine(start, dt.time()).timestamp())
    p2 = int(dt.datetime.combine(end + dt.timedelta(days=1), dt.time()).timestamp())

    for cand in yahoo_candidates(symbol):
        qs = urllib.parse.urlencode({"interval": "1d", "period1": p1, "period2": p2, "includePrePost": "false", "events": "div,splits"})
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(cand)}?{qs}"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=20) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
            bars = parse_yahoo_payload(payload)
            if bars:
                cache_file.write_text(json.dumps(payload), encoding="utf-8")
                return bars
        except Exception:
            time.sleep(0.15)
            continue
    return []


def safe_mean(xs: Sequence[float]) -> float:
    return sum(xs) / len(xs) if xs else 0.0


def safe_stdev(xs: Sequence[float]) -> float:
    return statistics.pstdev(xs) if len(xs) > 1 else 0.0


def compute_features(bars: Sequence[Bar], signal_date: dt.date) -> Optional[Dict[str, float]]:
    idx = -1
    for i, b in enumerate(bars):
        if b.date <= signal_date:
            idx = i
        else:
            break
    if idx < 260:
        return None

    closes = [b.close for b in bars]
    highs = [b.high for b in bars]
    lows = [b.low for b in bars]
    vols = [b.volume for b in bars]

    c = closes[idx]
    sma200 = safe_mean(closes[idx - 199 : idx + 1])
    sma200_prev20 = safe_mean(closes[idx - 219 : idx - 19])
    sma200_slope = (sma200 - sma200_prev20) / sma200_prev20 if sma200_prev20 else 0.0

    avg_vol20 = safe_mean(vols[idx - 19 : idx + 1])
    avg_dollar_vol20 = safe_mean([closes[j] * vols[j] for j in range(idx - 19, idx + 1)])
    rvol20 = vols[idx] / avg_vol20 if avg_vol20 else 0.0

    ret_4w = (c / closes[idx - 20] - 1.0) if closes[idx - 20] else 0.0
    ret_8w = (c / closes[idx - 40] - 1.0) if closes[idx - 40] else 0.0
    ret_12w = (c / closes[idx - 60] - 1.0) if closes[idx - 60] else 0.0

    high_52w = max(highs[idx - 251 : idx + 1])
    dist_52w_high = (c / high_52w - 1.0) if high_52w else 0.0

    trs = []
    for j in range(idx - 19, idx + 1):
        prev = closes[j - 1]
        trs.append(max(highs[j] - lows[j], abs(highs[j] - prev), abs(lows[j] - prev)))
    atr_pct20 = safe_mean(trs) / c if c else 0.0

    rets20 = []
    for j in range(idx - 19, idx + 1):
        prev = closes[j - 1]
        rets20.append((closes[j] / prev - 1.0) if prev else 0.0)
    rets100 = []
    for j in range(idx - 99, idx + 1):
        prev = closes[j - 1]
        rets100.append((closes[j] / prev - 1.0) if prev else 0.0)

    std100 = safe_stdev(rets100)
    compression = (safe_stdev(rets20) / std100) if std100 else 0.0
    prior_20_high = max(highs[idx - 20 : idx])
    breakout_20 = (c / prior_20_high - 1.0) if prior_20_high else 0.0

    up_vol, down_vol = 0.0, 0.0
    for j in range(idx - 19, idx + 1):
        if closes[j] >= closes[j - 1]:
            up_vol += vols[j]
        else:
            down_vol += vols[j]
    up_down = up_vol / (down_vol if down_vol > 0 else 1.0)

    structure = 0.0
    structure += 1.0 if c > sma200 else 0.0
    structure += 1.0 if sma200_slope > 0 else 0.0
    structure += 1.0 if breakout_20 > 0 else 0.0
    structure += 1.0 if compression < 0.85 else 0.0
    structure += 1.0 if rvol20 > 1.2 else 0.0

    return {
        "close_gt_sma200": 1.0 if c > sma200 else 0.0,
        "sma200_slope_20d": sma200_slope,
        "avg_dollar_volume_20": math.log10(max(avg_dollar_vol20, 1.0)),
        "rvol_20": rvol20,
        "ret_4w": ret_4w,
        "ret_8w": ret_8w,
        "ret_12w": ret_12w,
        "dist_52w_high": dist_52w_high,
        "atr_pct_20": atr_pct20,
        "compression_20_100": compression,
        "breakout_20": breakout_20,
        "up_down_vol_ratio_20": up_down,
        "structure_score": structure,
    }


def sigmoid(x: float) -> float:
    if x >= 0:
        z = math.exp(-x)
        return 1.0 / (1.0 + z)
    z = math.exp(x)
    return z / (1.0 + z)


def fit_logreg(X: List[List[float]], y: List[int], steps: int = 1000, lr: float = 0.08, l2: float = 0.001) -> List[float]:
    if not X:
        return []
    n, d = len(X), len(X[0])
    w = [0.0] * d
    random.seed(7)
    for _ in range(steps):
        grad = [0.0] * d
        for i in range(n):
            p = sigmoid(sum(w[j] * X[i][j] for j in range(d)))
            err = p - y[i]
            for j in range(d):
                grad[j] += err * X[i][j]
        for j in range(d):
            grad[j] = grad[j] / n + l2 * w[j]
            w[j] -= lr * grad[j]
    return w


def prepare_matrix(samples: List[Sample], means: Optional[List[float]] = None, stds: Optional[List[float]] = None) -> Tuple[List[List[float]], List[int], List[float], List[float]]:
    raw = [[s.features[f] for f in FEATURES] for s in samples]
    y = [s.label for s in samples]
    if means is None or stds is None:
        cols = list(zip(*raw)) if raw else []
        means = [safe_mean(list(c)) for c in cols]
        stds = [safe_stdev(list(c)) for c in cols]
    X: List[List[float]] = []
    for row in raw:
        z = []
        for j, v in enumerate(row):
            sd = stds[j] if stds[j] > 1e-9 else 1.0
            z.append((v - means[j]) / sd)
        X.append([1.0] + z)
    return X, y, means, stds


def precision_at_k(scored: List[Tuple[float, Sample]], k: int) -> float:
    top = scored[:k]
    return (sum(1 for _, s in top if s.label == 1) / min(k, len(top))) if top else 0.0


def recall_at_k(scored: List[Tuple[float, Sample]], k: int) -> float:
    pos = sum(1 for _, s in scored if s.label == 1)
    return (sum(1 for _, s in scored[:k] if s.label == 1) / pos) if pos else 0.0


def average_precision(scored: List[Tuple[float, Sample]]) -> float:
    pos = sum(1 for _, s in scored if s.label == 1)
    if pos == 0:
        return 0.0
    hit, total = 0, 0.0
    for i, (_, s) in enumerate(scored, start=1):
        if s.label == 1:
            hit += 1
            total += hit / i
    return total / pos

def main() -> int:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: workbook not found: {xlsx_path}")
        return 2

    asof_date = dt.date.fromisoformat(args.asof_date)
    rows = read_sheet(xlsx_path, args.sheet, args.weeks)
    if not rows:
        print("ERROR: no usable rows found")
        return 2

    symbols = sorted({r.symbol for r in rows})
    start = asof_date - dt.timedelta(days=820)
    end = asof_date + dt.timedelta(days=2)

    bars_by_symbol: Dict[str, List[Bar]] = {}
    failed: List[str] = []

    print(f"Loading market data for {len(symbols)} symbols...")
    for i, sym in enumerate(symbols, start=1):
        bars = fetch_yahoo_bars(sym, start, end, Path(args.cache_dir))
        if bars:
            bars_by_symbol[sym] = bars
        else:
            failed.append(sym)
        if i % 50 == 0:
            print(f"  {i}/{len(symbols)} symbols")

    print(f"Data loaded for {len(bars_by_symbol)} symbols; failed {len(failed)}")

    samples: List[Sample] = []
    for r in rows:
        bars = bars_by_symbol.get(r.symbol)
        if not bars:
            continue
        signal_date = asof_date - dt.timedelta(days=7 * r.week)
        feats = compute_features(bars, signal_date)
        if feats is None:
            continue
        samples.append(Sample(r.week, r.symbol, r.label, r.ranking, feats))

    if not samples:
        print("ERROR: no samples with computed features")
        return 2

    weeks = sorted({s.week for s in samples}, reverse=True)
    per_week = {w: [s for s in samples if s.week == w] for w in weeks}

    fold_rows = []
    coef_accum = [0.0] * (len(FEATURES) + 1)
    coef_count = 0

    for w in weeks:
        train_weeks = [x for x in weeks if x > w]
        if len(train_weeks) < args.min_train_weeks:
            continue
        train = [s for tw in train_weeks for s in per_week.get(tw, [])]
        test = per_week.get(w, [])
        if not test:
            continue

        pos = sum(s.label for s in train)
        neg = len(train) - pos
        if pos == 0 or neg == 0:
            continue

        X_train, y_train, means, stds = prepare_matrix(train)
        X_test, _, _, _ = prepare_matrix(test, means, stds)
        weights = fit_logreg(X_train, y_train)

        scored: List[Tuple[float, Sample]] = []
        for i, s in enumerate(test):
            p = sigmoid(sum(weights[j] * X_test[i][j] for j in range(len(weights))))
            scored.append((p, s))
        scored.sort(key=lambda t: t[0], reverse=True)

        p_at_k = precision_at_k(scored, args.top_k)
        r_at_k = recall_at_k(scored, args.top_k)
        ap = average_precision(scored)
        baseline = sum(s.label for s in test) / len(test)

        fold_rows.append({
            "week": w,
            "n_test": len(test),
            "positives": sum(s.label for s in test),
            "precision_at_k": round(p_at_k, 4),
            "recall_at_k": round(r_at_k, 4),
            "average_precision": round(ap, 4),
            "baseline_pos_rate": round(baseline, 4),
            "lift_precision_vs_baseline": round((p_at_k / baseline) if baseline > 0 else 0.0, 3),
        })

        for j, v in enumerate(weights):
            coef_accum[j] += v
        coef_count += 1

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    folds_csv = out_dir / "top5_walkforward_folds.csv"
    with folds_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["week", "n_test", "positives", "precision_at_k", "recall_at_k", "average_precision", "baseline_pos_rate", "lift_precision_vs_baseline"])
        writer.writeheader()
        writer.writerows(sorted(fold_rows, key=lambda x: x["week"], reverse=True))

    model_summary = []
    if coef_count > 0:
        avg_w = [x / coef_count for x in coef_accum]
        for i, feat in enumerate(FEATURES, start=1):
            model_summary.append((feat, avg_w[i], abs(avg_w[i])))
        model_summary.sort(key=lambda t: t[2], reverse=True)

    pos_samples = [s for s in samples if s.label == 1]
    neg_samples = [s for s in samples if s.label == 0]
    feature_lifts = []
    for feat in FEATURES:
        pvals = [s.features[feat] for s in pos_samples]
        nvals = [s.features[feat] for s in neg_samples]
        if pvals and nvals:
            pmean = safe_mean(pvals)
            nmean = safe_mean(nvals)
            pooled = (safe_stdev(pvals) + safe_stdev(nvals)) / 2.0
            effect = (pmean - nmean) / pooled if pooled > 1e-9 else 0.0
            feature_lifts.append((feat, pmean, nmean, effect))
    feature_lifts.sort(key=lambda t: abs(t[3]), reverse=True)

    report = out_dir / "top5_forecast_report.md"
    with report.open("w", encoding="utf-8") as f:
        f.write("# Top 5 Forecast Analysis\n\n")
        f.write(f"- As-of date: {asof_date.isoformat()}\n")
        f.write(f"- Workbook: {xlsx_path}\n")
        f.write(f"- Weeks analyzed: {args.weeks}\n")
        f.write(f"- Rows parsed: {len(rows)}\n")
        f.write(f"- Samples with features: {len(samples)}\n")
        f.write(f"- Symbols with price history: {len(bars_by_symbol)}\n")
        f.write(f"- Symbols failed: {len(failed)}\n\n")

        if fold_rows:
            avg_p = safe_mean([r["precision_at_k"] for r in fold_rows])
            avg_r = safe_mean([r["recall_at_k"] for r in fold_rows])
            avg_ap = safe_mean([r["average_precision"] for r in fold_rows])
            avg_lift = safe_mean([r["lift_precision_vs_baseline"] for r in fold_rows])
            f.write("## Walk-forward metrics\n\n")
            f.write(f"- Mean Precision@{args.top_k}: {avg_p:.3f}\n")
            f.write(f"- Mean Recall@{args.top_k}: {avg_r:.3f}\n")
            f.write(f"- Mean Average Precision: {avg_ap:.3f}\n")
            f.write(f"- Mean Precision Lift vs random baseline: {avg_lift:.2f}x\n\n")

        if model_summary:
            f.write("## Most predictive model features (avg logistic coefficient)\n\n")
            for feat, coeff, _ in model_summary[:8]:
                f.write(f"- `{feat}`: coeff={coeff:.4f}\n")
            f.write("\n")

        if feature_lifts:
            f.write("## Positive vs non-positive separation (effect size)\n\n")
            for feat, pmean, nmean, eff in feature_lifts[:10]:
                f.write(f"- `{feat}`: pos_mean={pmean:.4f}, nonpos_mean={nmean:.4f}, effect={eff:.3f}\n")
            f.write("\n")

        f.write("## Suggested forecasting checklist\n\n")
        f.write("1. Trend gate: close above 200MA and 200MA slope > 0.\n")
        f.write("2. Structure gate: breakout_20 > 0 and near 52-week highs.\n")
        f.write("3. Volume confirmation: RVOL20 > 1.2 and up/down volume ratio > 1.\n")
        f.write("4. Momentum confirmation: positive 8-week and 12-week returns.\n")
        f.write("5. Liquidity floor: log10(avg dollar volume 20d) above your tradability floor.\n")

    if failed:
        (out_dir / "top5_failed_symbols.txt").write_text("\n".join(failed), encoding="utf-8")

    print(f"Wrote: {folds_csv}")
    print(f"Wrote: {report}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
